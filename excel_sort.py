import argparse
import sys

try:
    import pandas as pd
except ImportError:
    sys.exit(
        "pandas library is required. Please install pandas and openpyxl before running this script."
    )

try:
    from openpyxl.styles import PatternFill
except ImportError:
    sys.exit(
        "openpyxl library is required. Please install pandas and openpyxl before running this script."
    )


def main() -> None:
    """Read an Excel file and sort it by a chosen column while keeping the first column untouched."""
    parser = argparse.ArgumentParser(
        description="Read an Excel file and sort it by a specific column"
    )
    parser.add_argument("input", help="Path to the Excel file to read")
    parser.add_argument("--column", required=True, help="Column name to sort by")
    parser.add_argument(
        "--output",
        help="Output Excel file path. If omitted, sorted data is printed",
    )
    args = parser.parse_args()

    df = pd.read_excel(args.input)
    df_original = df.copy()
    if args.column not in df.columns:
        sys.exit(f"Column '{args.column}' not found in the spreadsheet")
    if args.column == df.columns[0]:
        sys.exit(
            "Sorting by the first column is not allowed because the first column must remain unchanged"
        )

    first_col = df.columns[0]
    first_data = df[first_col]

    other_cols = df.columns[1:]
    df_other = df[other_cols].sort_values(by=args.column).reset_index(drop=True)
    df_sorted = pd.concat([first_data, df_other], axis=1)

    if args.output:
        with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
            df_sorted.to_excel(writer, index=False)
            sheet = writer.sheets["Sheet1"]
            highlight = PatternFill(start_color="CCFF99", end_color="CCFF99", fill_type="solid")
            for r in range(df_sorted.shape[0]):
                for c_idx, col in enumerate(df_sorted.columns[1:], start=2):
                    if df_sorted.iloc[r][col] != df_original.iloc[r][col]:
                        sheet.cell(row=r + 2, column=c_idx).fill = highlight
        print(f"Sorted data saved to {args.output}")
    else:
        print(df_sorted)


if __name__ == "__main__":
    main()
